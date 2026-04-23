"""
Módulo de manipulação de arquivos Excel.
Gerencia exportação e importação de pacientes e formulários em formato Excel.

SEGURANÇA:
- Não exporta nem importa dados sensíveis de identificação (CPF, RG, etc.)
- Exportação filtra apenas dados do usuário logado (feito no router)
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import logging
from pathlib import Path

from app.config import EXPORT_DIR

logger = logging.getLogger(__name__)

# Configuração de cores e formatação
HEADER_COLOR = "4472C4"  # Azul
HEADER_TEXT_COLOR = "FFFFFF"  # Branco

SURVEY_EXPORT_COLUMNS = [
    "Carimbo de data/hora",
    "Está em uso de terapia hormonizadora?",
    "Qual? (para homens trans)",
    "Qual? (para mulheres)",
    "Se já fez uso e não faz mais, qual o motivo da suspensão?",
    "Nome do paciente:",
    "Nome social: ",
    "Prontuário: ",
    "Telefone: ",
    "Raça: ",
    "Data de nascimento:",
    "TCLE assinado?",
    "Grau de escolaridade:",
    "Estado civil:",
    "Ocupação principal: ",
    "Renda familiar (considerar as pessoas com quem você mora): ",
    "Número de pessoas que vivem na casa (contando com você):",
    "Pratica atividade física?",
    "Se sim, em qual local?",
    "Quantas vezes por semana?",
    "Qual a duração média aproximada de cada sessão?",
    "Tipo de atividade física principal?",
    "Você consome bebidas alcoólicas?",
    "Se sim, qual a frequência do consumo de bebida alcoólica?",
    "Qual o tipo de bebida que consome com mais frequência?",
    "Você fuma?",
    "Qual tipo de cigarro você usa com mais frequência?",
    "Se fuma cigarro de maço, qual quantidade de cigarros fuma aproximadamente por dia?",
    "Você faz uso de drogas ilícitas?",
    "Se sim, qual/quais:",
    "Qual a frequência do uso de drogas ilícitas?",
    "Você sabe se já teve/tem alguma dessas doenças:",
    "História familiar de doenças (cardiovasculares, trombóticas, oncológicas, ou outros)?",
    "Medicamentos em uso atual (escrever medicamento, dose e frequência):",
    "PA",
    "Altura (cm):",
    "Peso (kg):",
    "IMC:",
    "Se sim, há quanto tempo?",
    "Você faz ou já fez algum acompanhamento de saúde mental?",
    "Você tem ou já teve algum diagnóstico relacionado a saúde mental?",
    "Caso a resposta para a pergunta anterior seja positiva, o quadro apresentou correlação com uso da hormonioterapia?",
    "Quais os seus objetivos quando começou/ para começar a hormonioterapia (homens trans)?",
    "Quais os seus objetivos quando começou/ para começar a hormonioterapia (mulheres trans)?",
    "Já teve alguma reação alérgica grave desde o início da hormonioterapia (anafilaxia)?",
    "Teve acne desde o início da hormonioterapia?",
    "Qual a região de acometimento da acne?",
    "Notou as áreas do pescoço, axilas ou outras áreas de dobra mais enegrecidas (acantose nigricans)?",
    "Notou o aparecimento de nódulos ou tumores palpáveis nas mamas?",
    "Com o uso da hormonioterapia, notou aumento de irritabilidade e/ou agressividade?",
    "Com o uso da hormonioterapia, notou aumento de estresse?",
    "Notou alguma outra mudança no humor?",
    "Como você avalia a sua autoestima durante o uso da hormonioterapia?",
    "Qual seu grau de satisfação com os resultados físicos/estéticos da hormonioterapia?",
    "Notou o aparecimento de outras condições de saúde (ex. nódulos hepáticos, outros nódulos, condições dermatológicas etc) ",
    "NÃO USAR ESSA PERGUNTA (DEIXEI SO PRA TRANSFERIR AS VELHAS PRA ELAS SEPARADAS)\nBioimpedância: ",
    "Glicose (data): ",
    "Perfil hepático (data):",
    "Perfil lipídico (data):",
    "Hemograma (data):",
    "Outros exames (data):",
    "TGO (data)",
    "TGP (data)",
    "Triglicerídeos (data)",
    "Colesterol Total (data)",
    "HDL (data)",
    "LDL (data)",
    "Testesterona (data)",
    "Hemoglobina (data)",
    "Hematócrito (data)",
    "Leucócitos  (data)",
    "Plaquetas (data)",
    "Bioimpedância - Peso (Kg)",
    "Bioimpedância - Músculo (Kg)",
    "Bioimpedância - Gordura (%)",
]

FIELD_TO_EXPORT_COLUMN = {
    "using_hormone_therapy": "Está em uso de terapia hormonizadora?",
    "hormone_type_men": "Qual? (para homens trans)",
    "hormone_type_women": "Qual? (para mulheres)",
    "suspension_reason": "Se já fez uso e não faz mais, qual o motivo da suspensão?",
    "patient_name": "Nome do paciente:",
    "social_name": "Nome social: ",
    "medical_record": "Prontuário: ",
    "phone": "Telefone: ",
    "race": "Raça: ",
    "birth_date": "Data de nascimento:",
    "tcle_signed": "TCLE assinado?",
    "education_level": "Grau de escolaridade:",
    "marital_status": "Estado civil:",
    "occupation": "Ocupação principal: ",
    "family_income": "Renda familiar (considerar as pessoas com quem você mora): ",
    "household_size": "Número de pessoas que vivem na casa (contando com você):",
    "physical_activity": "Pratica atividade física?",
    "physical_activity_location": "Se sim, em qual local?",
    "physical_activity_frequency": "Quantas vezes por semana?",
    "physical_activity_duration": "Qual a duração média aproximada de cada sessão?",
    "physical_activity_type": "Tipo de atividade física principal?",
    "alcohol_consumption": "Você consome bebidas alcoólicas?",
    "alcohol_frequency": "Se sim, qual a frequência do consumo de bebida alcoólica?",
    "alcohol_type": "Qual o tipo de bebida que consome com mais frequência?",
    "smoking": "Você fuma?",
    "cigarette_type": "Qual tipo de cigarro você usa com mais frequência?",
    "cigarettes_per_day": "Se fuma cigarro de maço, qual quantidade de cigarros fuma aproximadamente por dia?",
    "illicit_drugs": "Você faz uso de drogas ilícitas?",
    "illicit_drugs_type": "Se sim, qual/quais:",
    "illicit_drugs_frequency": "Qual a frequência do uso de drogas ilícitas?",
    "previous_diseases": "Você sabe se já teve/tem alguma dessas doenças:",
    "family_history": "História familiar de doenças (cardiovasculares, trombóticas, oncológicas, ou outros)?",
    "current_medications": "Medicamentos em uso atual (escrever medicamento, dose e frequência):",
    "blood_pressure": "PA",
    "height": "Altura (cm):",
    "weight": "Peso (kg):",
    "bmi": "IMC:",
    "mental_health_follow_up": "Você faz ou já fez algum acompanhamento de saúde mental?",
    "mental_health_diagnosis": "Você tem ou já teve algum diagnóstico relacionado a saúde mental?",
    "hormone_mental_health_correlation": "Caso a resposta para a pergunta anterior seja positiva, o quadro apresentou correlação com uso da hormonioterapia?",
    "hormone_objectives_men": "Quais os seus objetivos quando começou/ para começar a hormonioterapia (homens trans)?",
    "hormone_objectives_women": "Quais os seus objetivos quando começou/ para começar a hormonioterapia (mulheres trans)?",
    "allergic_reaction": "Já teve alguma reação alérgica grave desde o início da hormonioterapia (anafilaxia)?",
    "acne_occurrence": "Teve acne desde o início da hormonioterapia?",
    "acne_location": "Qual a região de acometimento da acne?",
    "acanthosis_nigricans": "Notou as áreas do pescoço, axilas ou outras áreas de dobra mais enegrecidas (acantose nigricans)?",
    "breast_nodules": "Notou o aparecimento de nódulos ou tumores palpáveis nas mamas?",
    "irritability_aggressiveness": "Com o uso da hormonioterapia, notou aumento de irritabilidade e/ou agressividade?",
    "stress_increase": "Com o uso da hormonioterapia, notou aumento de estresse?",
    "mood_changes": "Notou alguma outra mudança no humor?",
    "self_esteem": "Como você avalia a sua autoestima durante o uso da hormonioterapia?",
    "satisfaction_results": "Qual seu grau de satisfação com os resultados físicos/estéticos da hormonioterapia?",
    "other_health_conditions": "Notou o aparecimento de outras condições de saúde (ex. nódulos hepáticos, outros nódulos, condições dermatológicas etc) ",
    "glucose": "Glicose (data): ",
    "other_exams": "Outros exames (data):",
    "tgo": "TGO (data)",
    "tgp": "TGP (data)",
    "triglycerides": "Triglicerídeos (data)",
    "total_cholesterol": "Colesterol Total (data)",
    "hdl": "HDL (data)",
    "ldl": "LDL (data)",
    "testosterone": "Testesterona (data)",
    "hemoglobin": "Hemoglobina (data)",
    "hematocrit": "Hematócrito (data)",
    "leukocytes": "Leucócitos  (data)",
    "platelets": "Plaquetas (data)",
    "bioimpedance_weight": "Bioimpedância - Peso (Kg)",
    "bioimpedance_muscle": "Bioimpedância - Músculo (Kg)",
    "bioimpedance_fat": "Bioimpedância - Gordura (%)",
}

LEGACY_EXPORT_FIELDS = {
    "Perfil hepático (data):": "Perfil hepático (data):",
    "Perfil lipídico (data):": "Perfil lipídico (data):",
    "Hemograma (data):": "Hemograma (data):",
    "Bioimpedância:": "NÃO USAR ESSA PERGUNTA (DEIXEI SO PRA TRANSFERIR AS VELHAS PRA ELAS SEPARADAS)\nBioimpedância: ",
}

CARRY_FORWARD_FIELDS = {
    "using_hormone_therapy",
    "hormone_type_men",
    "hormone_type_women",
    "suspension_reason",
    "patient_name",
    "social_name",
    "medical_record",
    "phone",
    "race",
    "birth_date",
    "tcle_signed",
    "education_level",
    "marital_status",
    "occupation",
    "family_income",
    "household_size",
    "physical_activity",
    "physical_activity_location",
    "physical_activity_frequency",
    "physical_activity_duration",
    "physical_activity_type",
    "alcohol_consumption",
    "alcohol_frequency",
    "alcohol_type",
    "smoking",
    "cigarette_type",
    "cigarettes_per_day",
    "illicit_drugs",
    "illicit_drugs_type",
    "illicit_drugs_frequency",
    "previous_diseases",
    "family_history",
    "current_medications",
    "blood_pressure",
    "height",
    "weight",
    "bmi",
    "hormone_therapy_over_one_year",
}


def _clean_form_payload(form_data: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(form_data, dict):
        return {}
    return {
        key: value
        for key, value in form_data.items()
        if not str(key).startswith("_")
    }


def _get_legacy_fields(form_data: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    if not isinstance(form_data, dict):
        return {}
    legacy_fields = form_data.get("_legacy_fields")
    return legacy_fields if isinstance(legacy_fields, dict) else {}


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, (list, tuple, set)):
        return len(value) == 0
    return False


def _normalize_export_value(value: Any) -> Any:
    if _is_empty_value(value):
        return ""
    if isinstance(value, list):
        value = ", ".join(str(item) for item in value if not _is_empty_value(item))
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return f"'{value}"
    if isinstance(value, bool):
        return "Sim" if value else "Não"
    return value


def _normalize_datetime_for_excel(value: Any) -> Any:
    if _is_empty_value(value):
        return ""
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo else value
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
            return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
        except ValueError:
            return value
    return value


def _extract_duration_value(data: Dict[str, Any], legacy_fields: Dict[str, Any]) -> str:
    legacy_duration = legacy_fields.get("therapy_duration_text")
    if not _is_empty_value(legacy_duration):
        return str(legacy_duration)

    answer = data.get("hormone_therapy_over_one_year")
    if answer == "Sim":
        return "Mais de 1 ano"
    if answer == "Não":
        return "Menos de 1 ano"
    return ""


def _build_survey_row(
    patient: Dict[str, Any],
    response: Dict[str, Any],
    carry_forward_data: Dict[str, Any],
) -> Dict[str, Any]:
    form_data = _clean_form_payload(response.get("form_data"))
    legacy_fields = _get_legacy_fields(response.get("form_data"))

    for key, value in form_data.items():
        if key in CARRY_FORWARD_FIELDS and not _is_empty_value(value):
            carry_forward_data[key] = value

    row = {column: "" for column in SURVEY_EXPORT_COLUMNS}
    row["Carimbo de data/hora"] = _normalize_datetime_for_excel(response.get("response_date"))

    current_snapshot = dict(carry_forward_data)
    current_snapshot.update({key: value for key, value in form_data.items() if not _is_empty_value(value)})

    if _is_empty_value(current_snapshot.get("patient_name")):
        current_snapshot["patient_name"] = patient.get("full_name", "")

    for field_name, column_name in FIELD_TO_EXPORT_COLUMN.items():
        if field_name in CARRY_FORWARD_FIELDS:
            value = current_snapshot.get(field_name)
        else:
            value = form_data.get(field_name)
        row[column_name] = _normalize_export_value(value)

    for legacy_name, column_name in LEGACY_EXPORT_FIELDS.items():
        row[column_name] = _normalize_export_value(legacy_fields.get(legacy_name))

    row["Se sim, há quanto tempo?"] = _extract_duration_value(current_snapshot, legacy_fields)

    if not row["Nome do paciente:"]:
        row["Nome do paciente:"] = patient.get("full_name", "")

    return row


def ensure_export_dir():
    """
    Garante que o diretório de exportação existe.
    """
    Path(EXPORT_DIR).mkdir(parents=True, exist_ok=True)


def exportar_pacientes_excel(
    pacientes_data: List[Dict[str, Any]]
) -> str:
    """
    Exporta lista de pacientes com suas respostas de formulário para Excel.
    
    Cada paciente pode ter múltiplas respostas de formulário. Cada pergunta
    do formulário será uma coluna separada no Excel.
    
    Args:
        pacientes_data: Lista de dicionários contendo:
            - patient: dados do paciente (id, full_name, created_at)
            - form_responses: lista de respostas de formulário do paciente
    
    Returns:
        Caminho do arquivo Excel gerado
        
    Raises:
        Exception: Se houver erro ao gerar o arquivo
    """
    try:
        ensure_export_dir()

        rows = []
        for paciente_info in pacientes_data:
            patient = paciente_info.get("patient", {})
            form_responses = paciente_info.get("form_responses", [])

            if not form_responses:
                continue

            sorted_responses = sorted(
                form_responses,
                key=lambda item: item.get("response_date") or ""
            )

            carry_forward_data: Dict[str, Any] = {}
            for response in sorted_responses:
                rows.append(_build_survey_row(patient, response, carry_forward_data))

        if not rows:
            raise ValueError("Nenhum dado para exportar")

        df = pd.DataFrame(rows, columns=SURVEY_EXPORT_COLUMNS)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"pesquisa_export_{timestamp}.xlsx"
        filepath = str(EXPORT_DIR / filename)

        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Pacientes', index=False)
            worksheet = writer.sheets['Pacientes']

            header_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
            header_font = Font(bold=True, color=HEADER_TEXT_COLOR, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            worksheet.freeze_panes = "A2"

            data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = data_alignment

            for cell in worksheet["A"][1:]:
                if isinstance(cell.value, datetime):
                    cell.number_format = "dd/mm/yyyy hh:mm:ss"
            for cell in worksheet["K"][1:]:
                if isinstance(cell.value, datetime):
                    cell.number_format = "dd/mm/yyyy"

        logger.info(f"Arquivo Excel exportado com sucesso: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Erro ao exportar Excel: {str(e)}", exc_info=True)
        raise Exception(f"Erro ao exportar arquivo Excel: {str(e)}")


def importar_pacientes_excel(
    filepath: str
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Importa pacientes de um arquivo Excel.
    
    Valida a estrutura do arquivo, limpa os dados e retorna lista de pacientes
    validados ou lista de erros encontrados.
    
    Args:
        filepath: Caminho do arquivo Excel a ser importado
    
    Returns:
        Tupla contendo:
            - Lista de pacientes validados (dicionários)
            - Lista de erros encontrados (strings vazia se não houver erros)
    
    Raises:
        Exception: Se houver erro crítico ao processar o arquivo
    """
    erros = []
    pacientes_validados = []
    
    try:
        # Verifica se o arquivo existe
        filepath_obj = Path(filepath)
        if not filepath_obj.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {filepath}")
        
        # Lê o arquivo Excel
        try:
            df = pd.read_excel(filepath, engine='openpyxl')
        except Exception as e:
            raise Exception(f"Erro ao ler arquivo Excel: {str(e)}")
        
        # Valida colunas obrigatórias
        colunas_obrigatorias = ["Nome_Completo"]
        colunas_faltando = [col for col in colunas_obrigatorias if col not in df.columns]
        
        if colunas_faltando:
            erros.append(f"Colunas obrigatórias faltando: {', '.join(colunas_faltando)}")
            return pacientes_validados, erros
        
        # Processa cada linha
        for index, row in df.iterrows():
            linha_num = index + 2  # +2 porque index começa em 0 e há cabeçalho
            
            try:
                # Limpa e valida nome completo
                nome_completo = str(row.get("Nome_Completo", "")).strip()
                if not nome_completo or nome_completo.lower() in ["nan", "none", ""]:
                    erros.append(f"Linha {linha_num}: Nome completo é obrigatório")
                    continue
                
                if len(nome_completo) > 255:
                    erros.append(f"Linha {linha_num}: Nome completo muito longo (máximo 255 caracteres)")
                    continue
                
                # Prepara dados do paciente (sem CPF)
                paciente = {
                    "full_name": nome_completo
                }
                
                # Adiciona campos opcionais se existirem
                if "ID_Paciente" in df.columns and pd.notna(row.get("ID_Paciente")):
                    try:
                        paciente["id"] = int(row.get("ID_Paciente"))
                    except:
                        pass  # ID é opcional na importação
                
                pacientes_validados.append(paciente)
                
            except Exception as e:
                erros.append(f"Linha {linha_num}: Erro ao processar - {str(e)}")
                continue
        
        if not pacientes_validados and not erros:
            erros.append("Nenhum paciente válido encontrado no arquivo")
        
        logger.info(f"Importação concluída: {len(pacientes_validados)} pacientes válidos, {len(erros)} erros")
        return pacientes_validados, erros
        
    except FileNotFoundError as e:
        logger.error(f"Arquivo não encontrado: {str(e)}")
        raise
    except Exception as e:
        logger.error(f"Erro ao importar Excel: {str(e)}", exc_info=True)
        raise Exception(f"Erro ao importar arquivo Excel: {str(e)}")


def validar_estrutura_excel(filepath: str) -> Tuple[bool, List[str]]:
    """
    Valida a estrutura básica de um arquivo Excel antes da importação.
    
    Args:
        filepath: Caminho do arquivo Excel
    
    Returns:
        Tupla contendo:
            - True se a estrutura é válida, False caso contrário
            - Lista de mensagens de erro
    """
    erros = []
    
    try:
        # Verifica extensão
        if not str(filepath).lower().endswith(('.xlsx', '.xls')):
            erros.append("Arquivo deve ser do tipo .xlsx ou .xls")
            return False, erros
        
        # Tenta abrir o arquivo
        try:
            wb = load_workbook(filepath, read_only=True)
            if len(wb.sheetnames) == 0:
                erros.append("Arquivo não contém planilhas")
                return False, erros
            wb.close()
        except Exception as e:
            erros.append(f"Arquivo Excel corrompido ou inválido: {str(e)}")
            return False, erros
        
        return True, erros
        
    except Exception as e:
        erros.append(f"Erro ao validar arquivo: {str(e)}")
        return False, erros
